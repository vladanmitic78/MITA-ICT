from fastapi import APIRouter, HTTPException, Depends, status
from fastapi.responses import StreamingResponse
from datetime import datetime
from typing import List
import logging
import io
import uuid

from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from models import (
    Service, ServiceCreate, ServiceUpdate,
    SaasProduct, SaasProductCreate, SaasProductUpdate,
    AboutContent, AboutContentUpdate,
    Contact, ContactUpdate,
    SocialIntegrations
)
from auth import get_current_user

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/admin", tags=["Admin CRUD"])


def get_db():
    """Dependency to get database - will be set by main app"""
    from server import db
    return db


# ==================== SERVICES MANAGEMENT ====================

@router.post("/services", response_model=Service)
async def create_service(service: ServiceCreate, current_user: dict = Depends(get_current_user)):
    """Create a new service (admin only)"""
    db = get_db()
    service_data = Service(
        title=service.title,
        description=service.description,
        icon=service.icon
    )
    await db.services.insert_one(service_data.dict())
    logger.info(f"✅ Service created: {service.title}")
    return service_data


@router.put("/services/{service_id}", response_model=Service)
async def update_service(
    service_id: str,
    service: ServiceUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update a service (admin only)"""
    db = get_db()
    existing = await db.services.find_one({"id": service_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Service not found")
    
    update_data = service.dict(exclude_unset=True)
    update_data['updated_at'] = datetime.utcnow()
    
    await db.services.update_one({"id": service_id}, {"$set": update_data})
    updated = await db.services.find_one({"id": service_id})
    logger.info(f"✅ Service updated: {service_id}")
    return Service(**updated)


@router.delete("/services/{service_id}")
async def delete_service(service_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a service (admin only)"""
    db = get_db()
    result = await db.services.delete_one({"id": service_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Service not found")
    logger.info(f"✅ Service deleted: {service_id}")
    return {"message": "Service deleted successfully"}


# ==================== SAAS PRODUCTS MANAGEMENT ====================

@router.post("/saas-products", response_model=SaasProduct)
async def create_saas_product(product: SaasProductCreate, current_user: dict = Depends(get_current_user)):
    """Create a new SaaS product (admin only)"""
    db = get_db()
    product_data = SaasProduct(
        name=product.name,
        description=product.description,
        url=product.url,
        features=product.features
    )
    await db.saas_products.insert_one(product_data.dict())
    logger.info(f"✅ SaaS product created: {product.name}")
    return product_data


@router.put("/saas-products/{product_id}", response_model=SaasProduct)
async def update_saas_product(
    product_id: str,
    product: SaasProductUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update a SaaS product (admin only)"""
    db = get_db()
    existing = await db.saas_products.find_one({"id": product_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Product not found")
    
    update_data = product.dict(exclude_unset=True)
    update_data['updated_at'] = datetime.utcnow()
    
    await db.saas_products.update_one({"id": product_id}, {"$set": update_data})
    updated = await db.saas_products.find_one({"id": product_id})
    logger.info(f"✅ SaaS product updated: {product_id}")
    return SaasProduct(**updated)


@router.delete("/saas-products/{product_id}")
async def delete_saas_product(product_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a SaaS product (admin only)"""
    db = get_db()
    result = await db.saas_products.delete_one({"id": product_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    logger.info(f"✅ SaaS product deleted: {product_id}")
    return {"message": "Product deleted successfully"}


# ==================== ABOUT CONTENT MANAGEMENT ====================

@router.put("/about", response_model=AboutContent)
async def update_about_content(about: AboutContentUpdate, current_user: dict = Depends(get_current_user)):
    """Update About page content (admin only)"""
    db = get_db()
    existing = await db.about_content.find_one()
    
    update_data = about.dict(exclude_unset=True)
    update_data['updated_at'] = datetime.utcnow()
    
    if existing:
        update_data['id'] = existing['id']
        await db.about_content.update_one({"id": existing['id']}, {"$set": update_data})
    else:
        new_about = AboutContent(**update_data)
        await db.about_content.insert_one(new_about.dict())
    
    updated = await db.about_content.find_one()
    logger.info("✅ About content updated")
    return AboutContent(**updated)


# ==================== CONTACTS MANAGEMENT ====================

@router.get("/contacts")
async def get_contacts(
    current_user: dict = Depends(get_current_user),
    page: int = 1,
    limit: int = 50,
    search: str = None
):
    """Get contacts with pagination (admin only)"""
    db = get_db()
    
    # Build query filter
    query = {}
    if search:
        query = {
            "$or": [
                {"name": {"$regex": search, "$options": "i"}},
                {"email": {"$regex": search, "$options": "i"}},
                {"phone": {"$regex": search, "$options": "i"}},
                {"service": {"$regex": search, "$options": "i"}}
            ]
        }
    
    # Get total count for pagination
    total = await db.contacts.count_documents(query)
    
    # Calculate skip value
    skip = (page - 1) * limit
    
    # Use projection to fetch only needed fields
    projection = {"_id": 0, "id": 1, "name": 1, "email": 1, "phone": 1, "service": 1, "comment": 1, "created_at": 1}
    
    contacts = await db.contacts.find(query, projection).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    return {
        "data": contacts,
        "total": total,
        "page": page,
        "limit": limit,
        "pages": (total + limit - 1) // limit
    }


@router.put("/contacts/{contact_id}")
async def update_contact(
    contact_id: str,
    contact: ContactUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update a contact (admin only)"""
    db = get_db()
    existing = await db.contacts.find_one({"id": contact_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Contact not found")
    
    update_data = contact.dict(exclude_unset=True)
    
    await db.contacts.update_one({"id": contact_id}, {"$set": update_data})
    updated = await db.contacts.find_one({"id": contact_id})
    
    logger.info(f"✅ Contact updated: {contact_id}")
    return {
        "id": updated.get("id"),
        "name": updated.get("name"),
        "email": updated.get("email"),
        "phone": updated.get("phone"),
        "service": updated.get("service"),
        "comment": updated.get("comment"),
        "created_at": updated.get("created_at")
    }


@router.delete("/contacts/{contact_id}")
async def delete_contact(contact_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a contact (admin only)"""
    db = get_db()
    result = await db.contacts.delete_one({"id": contact_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Contact not found")
    logger.info(f"✅ Contact deleted: {contact_id}")
    return {"message": "Contact deleted successfully"}


@router.get("/contacts/export-pdf")
async def export_contacts_pdf(current_user: dict = Depends(get_current_user)):
    """Export contacts as PDF (admin only)"""
    db = get_db()
    contacts = await db.contacts.find().sort("created_at", -1).to_list(1000)
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=20,
        textColor=colors.HexColor('#00FFD1')
    )
    elements.append(Paragraph("MITA ICT - Contact Submissions", title_style))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    table_data = [['Name', 'Email', 'Phone', 'Service', 'Date']]
    
    for contact in contacts:
        table_data.append([
            contact.get('name', 'N/A'),
            contact.get('email', 'N/A'),
            contact.get('phone', 'N/A'),
            contact.get('service', 'N/A'),
            contact.get('created_at', datetime.now()).strftime('%Y-%m-%d') if contact.get('created_at') else 'N/A'
        ])
    
    table = Table(table_data, colWidths=[1.3*inch, 1.8*inch, 1*inch, 1.5*inch, 0.9*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#00FFD1')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#1a1a1a')),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.white),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#333333')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#1a1a1a'), colors.HexColor('#222222')]),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    
    logger.info("✅ Contacts PDF export generated")
    return StreamingResponse(
        buffer,
        media_type='application/pdf',
        headers={"Content-Disposition": f"attachment; filename=contacts_{datetime.now().strftime('%Y%m%d')}.pdf"}
    )


@router.get("/contacts/export-excel")
async def export_contacts_excel(current_user: dict = Depends(get_current_user)):
    """Export contacts as Excel (admin only)"""
    db = get_db()
    contacts = await db.contacts.find().sort("created_at", -1).to_list(1000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"
    
    headers = ['Name', 'Email', 'Phone', 'Service', 'Comment', 'Date']
    header_fill = PatternFill(start_color='00FFD1', end_color='00FFD1', fill_type='solid')
    header_font = Font(bold=True, color='000000')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    for row, contact in enumerate(contacts, 2):
        ws.cell(row=row, column=1, value=contact.get('name', ''))
        ws.cell(row=row, column=2, value=contact.get('email', ''))
        ws.cell(row=row, column=3, value=contact.get('phone', ''))
        ws.cell(row=row, column=4, value=contact.get('service', ''))
        ws.cell(row=row, column=5, value=contact.get('comment', ''))
        created = contact.get('created_at')
        ws.cell(row=row, column=6, value=created.strftime('%Y-%m-%d %H:%M') if created else '')
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 18
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    logger.info("✅ Contacts Excel export generated")
    return StreamingResponse(
        buffer,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": f"attachment; filename=contacts_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )


# ==================== SOCIAL MEDIA INTEGRATIONS ====================

@router.get("/social-integrations", response_model=SocialIntegrations)
async def get_social_integrations(current_user: dict = Depends(get_current_user)):
    """Get social media integrations (admin only)"""
    db = get_db()
    existing = await db.social_integrations.find_one()
    
    if not existing:
        default_integrations = SocialIntegrations()
        return default_integrations
    
    return SocialIntegrations(**existing)


@router.put("/social-integrations", response_model=SocialIntegrations)
async def update_social_integrations(
    integrations: SocialIntegrations,
    current_user: dict = Depends(get_current_user)
):
    """Update social media integrations (admin only)"""
    db = get_db()
    existing = await db.social_integrations.find_one()
    
    integration_dict = integrations.dict()
    integration_dict['updated_at'] = datetime.utcnow()
    
    if existing:
        integration_dict['id'] = existing['id']
        await db.social_integrations.update_one(
            {"id": existing['id']},
            {"$set": integration_dict}
        )
    else:
        await db.social_integrations.insert_one(integration_dict)
    
    logger.info("✅ Social media integrations updated")
    return SocialIntegrations(**integration_dict)
