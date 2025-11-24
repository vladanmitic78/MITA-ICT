from fastapi import FastAPI, APIRouter, HTTPException, Depends, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from motor.motor_asyncio import AsyncIOMotorClient
from dotenv import load_dotenv
from pathlib import Path
import os
import logging
from datetime import datetime
from typing import List
import httpx
import io
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from models import (
    Service, ServiceCreate, ServiceUpdate,
    SaasProduct, SaasProductCreate, SaasProductUpdate,
    Contact, ContactCreate, ContactUpdate,
    AboutContent, AboutContentUpdate,
    AdminLogin, Token, Admin, ChangePassword
)
from auth import (
    verify_password, get_password_hash, create_access_token, 
    get_current_user, init_admin_user
)
from email_service import send_contact_email, send_auto_response_email

# Setup
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ.get('MONGO_URL')
if not mongo_url:
    raise ValueError('MONGO_URL environment variable is required')

client = AsyncIOMotorClient(mongo_url)

db_name = os.environ.get('DB_NAME')
if not db_name:
    raise ValueError('DB_NAME environment variable is required')

db = client[db_name]

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Create the main app
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ==================== STARTUP & INITIALIZATION ====================

@app.on_event("startup")
async def startup_db_client():
    """Initialize database with default data"""
    logger.info("🚀 Starting MITA ICT Backend...")
    
    # Initialize admin user
    await init_admin_user(db)
    
    # Initialize default services if none exist
    services_count = await db.services.count_documents({})
    if services_count == 0:
        default_services = [
            {
                "id": "1",
                "title": "IT and Telecommunication",
                "description": "Comprehensive IT and telecom consulting services with 20+ years of industry experience. From infrastructure to advanced solutions.",
                "icon": "Network",
                "created_at": datetime.utcnow(),
                "updated_at": datetime.utcnow()
            },
            {
                "id": "2",
                "title": "Company Registration in Sweden",
                "description": "Complete support for company registration and business setup in Sweden. Navigate Swedish regulations with ease.",
                "icon": "Building2",
                "created_at": datetime.utcnow(),
                "updated_at": datetime.utcnow()
            },
            {
                "id": "3",
                "title": "Leading Teams",
                "description": "Expert leadership consulting for sales and engineering teams. Build high-performing organizations.",
                "icon": "Users",
                "created_at": datetime.utcnow(),
                "updated_at": datetime.utcnow()
            }
        ]
        await db.services.insert_many(default_services)
        logger.info("✅ Default services initialized")
    
    # Initialize default SaaS products if none exist
    saas_count = await db.saas_products.count_documents({})
    if saas_count == 0:
        default_products = [
            {
                "id": "1",
                "title": "MITACRM",
                "description": "Powerful CRM solution designed for modern businesses. Streamline your customer relationships.",
                "link": "https://mitacrm.com/",
                "features": ["Contact Management", "Sales Pipeline", "Analytics Dashboard", "Integration Ready"],
                "created_at": datetime.utcnow(),
                "updated_at": datetime.utcnow()
            },
            {
                "id": "2",
                "title": "Routing System",
                "description": "Advanced routing system for telecommunications and network management.",
                "link": "https://trustcode.dev/",
                "features": ["Smart Routing", "Real-time Monitoring", "Scalable Architecture", "API Access"],
                "created_at": datetime.utcnow(),
                "updated_at": datetime.utcnow()
            },
            {
                "id": "3",
                "title": "White Label Software",
                "description": "Customizable white label solutions for your business needs.",
                "link": "#",
                "features": ["Full Customization", "Your Branding", "Quick Deployment", "Ongoing Support"],
                "created_at": datetime.utcnow(),
                "updated_at": datetime.utcnow()
            }
        ]
        await db.saas_products.insert_many(default_products)
        logger.info("✅ Default SaaS products initialized")
    
    # Initialize default About content if none exists
    about_count = await db.about_content.count_documents({})
    if about_count == 0:
        default_about = {
            "id": "1",
            "title": "About MITA ICT",
            "content": "We bring over 20 years of distinguished experience in the IT and telecommunications industry. Our journey has been marked by successfully leading sales and engineering teams, implementing cutting-edge solutions, and driving organizational excellence.\n\nOur expertise spans across multiple domains including IT infrastructure, telecommunications networks, and enterprise software solutions. We have a proven track record in selling and implementing sophisticated software systems such as OSS (Operations Support Systems), OBS (Order and Billing Systems), and comprehensive cybersecurity solutions including EDR (Endpoint Detection and Response), MDR (Managed Detection and Response), and XDR (Extended Detection and Response).\n\nAt MITA ICT, client satisfaction is not just a goal—it's our foundation. We pride ourselves on understanding our clients' unique challenges and delivering tailored solutions that drive real business value. Our approach combines technical excellence with strategic thinking, ensuring that technology serves your business objectives.\n\nWhether you're looking to optimize your IT infrastructure, implement new telecommunications systems, or build high-performing teams, we bring the experience, expertise, and dedication to help you succeed.",
            "expertise": [
                {
                    "title": "IT Infrastructure",
                    "items": ["Network Design", "Cloud Solutions", "System Integration"]
                },
                {
                    "title": "Telecommunications",
                    "items": ["OSS Implementation", "Network Optimization", "Voice & Data Solutions"]
                },
                {
                    "title": "Cybersecurity",
                    "items": ["EDR/MDR/XDR Solutions", "Security Audits", "Compliance Management"]
                },
                {
                    "title": "Leadership",
                    "items": ["Team Building", "Sales Management", "P&L Optimization"]
                }
            ],
            "updated_at": datetime.utcnow()
        }
        await db.about_content.insert_one(default_about)
        logger.info("✅ Default About content initialized")
    
    logger.info("✅ Database initialized successfully")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
    logger.info("🔴 Database connection closed")

# ==================== PUBLIC ENDPOINTS ====================

@api_router.get("/")
async def root():
    return {"message": "MITA ICT API - Where Technology Meets Strategy"}

@api_router.get("/services", response_model=List[Service])
async def get_services():
    """Get all consulting services"""
    services = await db.services.find().to_list(1000)
    return [Service(**service) for service in services]

@api_router.get("/saas-products", response_model=List[SaasProduct])
async def get_saas_products():
    """Get all SaaS products"""
    products = await db.saas_products.find().to_list(1000)
    return [SaasProduct(**product) for product in products]

@api_router.get("/about", response_model=AboutContent)
async def get_about_content():
    """Get About page content"""
    about = await db.about_content.find_one()
    if not about:
        raise HTTPException(status_code=404, detail="About content not found")
    return AboutContent(**about)

@api_router.post("/contact", response_model=dict)
async def submit_contact(contact: ContactCreate):
    """Submit contact form"""
    try:
        # Verify reCAPTCHA token
        recaptcha_secret = os.environ.get('RECAPTCHA_SECRET_KEY', '')
        if recaptcha_secret and recaptcha_secret != 'YOUR_RECAPTCHA_SECRET_KEY':
            async with httpx.AsyncClient() as client_http:
                recaptcha_response = await client_http.post(
                    'https://www.google.com/recaptcha/api/siteverify',
                    data={
                        'secret': recaptcha_secret,
                        'response': contact.recaptcha_token
                    }
                )
                recaptcha_result = recaptcha_response.json()
                
                if not recaptcha_result.get('success'):
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="reCAPTCHA verification failed"
                    )
        
        # Save to database
        contact_data = Contact(
            name=contact.name,
            email=contact.email,
            phone=contact.phone,
            service=contact.service,
            comment=contact.comment
        )
        await db.contacts.insert_one(contact_data.dict())
        
        # Send email notifications
        try:
            # Send notification to admin
            await send_contact_email(
                name=contact.name,
                email=contact.email,
                phone=contact.phone,
                service=contact.service,
                comment=contact.comment
            )
            logger.info(f"✅ Admin notification email sent for: {contact.email}")
            
            # Send auto-response to user
            await send_auto_response_email(
                name=contact.name,
                email=contact.email,
                phone=contact.phone,
                service=contact.service
            )
            logger.info(f"✅ Auto-response email sent to: {contact.email}")
            
        except Exception as e:
            logger.error(f"❌ Email sending failed: {str(e)}")
            # Still save to database even if email fails
        
        return {
            "success": True,
            "message": "Thank you for contacting us. We will get back to you soon!"
        }
        
    except HTTPException as he:
        raise he
    except Exception as e:
        logger.error(f"❌ Contact form submission failed: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to submit contact form"
        )

# ==================== ADMIN AUTHENTICATION ====================

@api_router.post("/admin/login", response_model=Token)
async def admin_login(credentials: AdminLogin):
    """Admin login with username/password"""
    admin = await db.admins.find_one({"username": credentials.username})
    
    if not admin or not verify_password(credentials.password, admin["password_hash"]):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    access_token = create_access_token(data={"sub": admin["username"]})
    logger.info(f"✅ Admin logged in: {credentials.username}")
    
    return {"access_token": access_token, "token_type": "bearer"}

@api_router.post("/admin/google-login", response_model=Token)
async def admin_google_login():
    """Admin login with Google OAuth (mock for now)"""
    # For MVP, we'll create a token for Google login
    # In production, this would verify Google OAuth token
    access_token = create_access_token(data={"sub": "google_admin"})
    logger.info(f"✅ Admin logged in via Google")
    
    return {"access_token": access_token, "token_type": "bearer"}

@api_router.post("/admin/logout")
async def admin_logout(current_user: dict = Depends(get_current_user)):
    """Admin logout"""
    logger.info(f"✅ Admin logged out: {current_user['username']}")
    return {"message": "Successfully logged out"}

@api_router.post("/admin/change-password")
async def change_password(
    password_data: ChangePassword,
    current_user: dict = Depends(get_current_user)
):
    """Change admin password"""
    username = current_user['username']
    admin = await db.admins.find_one({"username": username})
    
    if not admin:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Admin user not found"
        )
    
    # Verify current password
    if not verify_password(password_data.current_password, admin["password_hash"]):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Current password is incorrect"
        )
    
    # Update password
    new_password_hash = get_password_hash(password_data.new_password)
    await db.admins.update_one(
        {"username": username},
        {"$set": {"password_hash": new_password_hash}}
    )
    
    logger.info(f"✅ Password changed for admin: {username}")
    return {"message": "Password updated successfully"}

# ==================== PROTECTED ADMIN ENDPOINTS ====================

@api_router.get("/admin/contacts", response_model=List[Contact])
async def get_contacts(current_user: dict = Depends(get_current_user)):
    """Get all contact submissions (admin only)"""
    contacts = await db.contacts.find().sort("created_at", -1).to_list(1000)
    return [Contact(**contact) for contact in contacts]

# Service Management
@api_router.post("/admin/services", response_model=Service)
async def create_service(
    service: ServiceCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create new service (admin only)"""
    service_data = Service(**service.dict())
    await db.services.insert_one(service_data.dict())
    logger.info(f"✅ Service created: {service_data.title}")
    return service_data

@api_router.put("/admin/services/{service_id}", response_model=Service)
async def update_service(
    service_id: str,
    service: ServiceUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update service (admin only)"""
    existing_service = await db.services.find_one({"id": service_id})
    if not existing_service:
        raise HTTPException(status_code=404, detail="Service not found")
    
    update_data = service.dict()
    update_data["updated_at"] = datetime.utcnow()
    
    await db.services.update_one(
        {"id": service_id},
        {"$set": update_data}
    )
    
    updated_service = await db.services.find_one({"id": service_id})
    logger.info(f"✅ Service updated: {service_id}")
    return Service(**updated_service)

@api_router.delete("/admin/services/{service_id}")
async def delete_service(
    service_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete service (admin only)"""
    result = await db.services.delete_one({"id": service_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Service not found")
    
    logger.info(f"✅ Service deleted: {service_id}")
    return {"message": "Service deleted successfully"}

# SaaS Product Management
@api_router.post("/admin/saas-products", response_model=SaasProduct)
async def create_saas_product(
    product: SaasProductCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create new SaaS product (admin only)"""
    product_data = SaasProduct(**product.dict())
    await db.saas_products.insert_one(product_data.dict())
    logger.info(f"✅ SaaS product created: {product_data.title}")
    return product_data

@api_router.put("/admin/saas-products/{product_id}", response_model=SaasProduct)
async def update_saas_product(
    product_id: str,
    product: SaasProductUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update SaaS product (admin only)"""
    existing_product = await db.saas_products.find_one({"id": product_id})
    if not existing_product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    update_data = product.dict()
    update_data["updated_at"] = datetime.utcnow()
    
    await db.saas_products.update_one(
        {"id": product_id},
        {"$set": update_data}
    )
    
    updated_product = await db.saas_products.find_one({"id": product_id})
    logger.info(f"✅ SaaS product updated: {product_id}")
    return SaasProduct(**updated_product)

@api_router.delete("/admin/saas-products/{product_id}")
async def delete_saas_product(
    product_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete SaaS product (admin only)"""
    result = await db.saas_products.delete_one({"id": product_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    
    logger.info(f"✅ SaaS product deleted: {product_id}")
    return {"message": "Product deleted successfully"}

# About Content Management
@api_router.put("/admin/about", response_model=AboutContent)
async def update_about_content(
    about: AboutContentUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update About page content (admin only)"""
    existing_about = await db.about_content.find_one()
    
    if not existing_about:
        # Create new about content if none exists
        about_data = AboutContent(
            title=about.title,
            content=about.content
        )
        await db.about_content.insert_one(about_data.dict())
        logger.info(f"✅ About content created")
        return about_data
    
    # Update existing content
    update_data = about.dict()
    update_data["updated_at"] = datetime.utcnow()
    
    await db.about_content.update_one(
        {"id": existing_about["id"]},
        {"$set": update_data}
    )
    
    updated_about = await db.about_content.find_one({"id": existing_about["id"]})
    logger.info(f"✅ About content updated")
    return AboutContent(**updated_about)

# Export Endpoints
@api_router.get("/admin/contacts/export/pdf")
async def export_contacts_pdf(current_user: dict = Depends(get_current_user)):
    """Export contacts to PDF (admin only)"""
    try:
        contacts = await db.contacts.find().sort("created_at", -1).to_list(1000)
        
        # Create PDF in memory
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=50, bottomMargin=30)
        
        # Container for the 'Flowable' objects
        elements = []
        
        # Add title
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#00D9FF'),
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        title = Paragraph("MITA ICT - Contact Submissions", title_style)
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        # Add export date
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            alignment=1
        )
        export_date = Paragraph(f"Exported on: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", date_style)
        elements.append(export_date)
        elements.append(Spacer(1, 20))
        
        if contacts:
            # Prepare table data
            table_data = [['Name', 'Email', 'Phone', 'Service', 'Date']]
            
            for contact in contacts:
                created_at = contact.get('created_at', datetime.utcnow())
                if isinstance(created_at, str):
                    created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                
                table_data.append([
                    contact.get('name', 'N/A'),
                    contact.get('email', 'N/A'),
                    contact.get('phone', 'N/A'),
                    contact.get('service', 'N/A'),
                    created_at.strftime('%Y-%m-%d')
                ])
            
            # Create table
            table = Table(table_data, colWidths=[1.5*inch, 2*inch, 1.5*inch, 1.5*inch, 1*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#00D9FF')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
            ]))
            elements.append(table)
            
            # Add details section for comments
            elements.append(Spacer(1, 30))
            details_title = Paragraph("<b>Contact Details with Comments:</b>", styles['Heading2'])
            elements.append(details_title)
            elements.append(Spacer(1, 12))
            
            for i, contact in enumerate(contacts, 1):
                detail_style = styles['Normal']
                detail_text = f"""
                <b>{i}. {contact.get('name', 'N/A')}</b><br/>
                Email: {contact.get('email', 'N/A')}<br/>
                Phone: {contact.get('phone', 'N/A')}<br/>
                Service: {contact.get('service', 'N/A')}<br/>
                Comment: {contact.get('comment', 'No comment provided')}<br/>
                Date: {created_at.strftime('%Y-%m-%d %H:%M')}
                """
                elements.append(Paragraph(detail_text, detail_style))
                elements.append(Spacer(1, 15))
        else:
            no_data = Paragraph("No contact submissions available.", styles['Normal'])
            elements.append(no_data)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        logger.info(f"✅ PDF export generated: {len(contacts)} contacts")
        
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=mita_contacts_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
            }
        )
        
    except Exception as e:
        logger.error(f"❌ PDF export failed: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to generate PDF: {str(e)}"
        )

@api_router.get("/admin/contacts/export/excel")
async def export_contacts_excel(current_user: dict = Depends(get_current_user)):
    """Export contacts to Excel (admin only)"""
    try:
        contacts = await db.contacts.find().sort("created_at", -1).to_list(1000)
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Contacts"
        
        # Add title
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = 'MITA ICT - Contact Submissions'
        title_cell.font = Font(size=16, bold=True, color="000000")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color="00D9FF", end_color="00D9FF", fill_type="solid")
        
        # Add export date
        ws.merge_cells('A2:F2')
        date_cell = ws['A2']
        date_cell.value = f"Exported on: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
        date_cell.font = Font(size=10, italic=True)
        date_cell.alignment = Alignment(horizontal='center')
        
        # Add headers
        headers = ['Name', 'Email', 'Phone', 'Service', 'Comment', 'Submitted Date']
        ws.append([])  # Empty row
        ws.append(headers)
        
        # Style headers
        header_row = ws[4]
        for cell in header_row:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data
        for contact in contacts:
            created_at = contact.get('created_at', datetime.utcnow())
            if isinstance(created_at, str):
                created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
            
            ws.append([
                contact.get('name', 'N/A'),
                contact.get('email', 'N/A'),
                contact.get('phone', 'N/A'),
                contact.get('service', 'N/A'),
                contact.get('comment', 'No comment provided'),
                created_at.strftime('%Y-%m-%d %H:%M')
            ])
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 20
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        logger.info(f"✅ Excel export generated: {len(contacts)} contacts")
        
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=mita_contacts_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
            }
        )
        
    except Exception as e:
        logger.error(f"❌ Excel export failed: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to generate Excel: {str(e)}"
        )


# Include the router in the main app
app.include_router(api_router)

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)
